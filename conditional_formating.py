from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule


def text_formula(search_for, range):
    return [f'NOT(ISERROR(SEARCH("{search_for}",{range[:2]})))']


def custom_conditional_formatting(ws, range, type, start=6, end=10, competence='N'):
    bgcolors = {'EKv': "C6EFCE", 'EKü': "C4D79B", 'GKv': "FFEB9C", 'GKü': "FFC000", 'N': "FFC7CE"}
    style_ekv = DifferentialStyle(font=Font(color="006100"), fill=PatternFill(bgColor=bgcolors['EKv']))
    style_eku = DifferentialStyle(font=Font(color="006100"), fill=PatternFill(bgColor=bgcolors['EKü']))
    style_gkv = DifferentialStyle(font=Font(color="9C5700"), fill=PatternFill(bgColor=bgcolors['GKv']))
    style_gku = DifferentialStyle(font=Font(color="9C5700"), fill=PatternFill(bgColor=bgcolors['GKü']))
    style_n = DifferentialStyle(font=Font(color="9C0006"), fill=PatternFill(bgColor=bgcolors['N']))

    if type == 'GEK':
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_ekv,
                                                  formula=text_formula('EKv', range)))
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_eku,
                                                  formula=text_formula('EKü', range)))
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_gkv,
                                                  formula=text_formula('GKv', range)))
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_gku,
                                                  formula=text_formula('GKü', range)))
    elif type == 'GK':
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_gkv,
                                                  formula=text_formula('v', range)))
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_gku,
                                                  formula=text_formula('ü', range)))
    elif type == 'EK':
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_ekv,
                                                  formula=text_formula('v', range)))
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_eku,
                                                  formula=text_formula('ü', range)))
    elif type == 'marks':
        ws.conditional_formatting.add(range, ColorScaleRule(start_type='num', start_value=1, start_color='63BE7B',
                                                            mid_type='num', mid_value=3, mid_color='FFEB84',
                                                            end_type='num', end_value=5, end_color='f8696b'))
    elif type == 'points':
        ws.conditional_formatting.add(range,
                                      CellIsRule(operator='lessThan', formula=[start], stopIfTrue=False,
                                                 fill=PatternFill(bgColor="F8696B")))
        ws.conditional_formatting.add(range,
                                      ColorScaleRule(start_type='num', start_value=start, start_color='FBAA77',
                                                     mid_type='percent', mid_value=70, mid_color='FFEB84',
                                                     end_type='num', end_value=end, end_color='63BE7B'))
    elif type == 'scale':
        ws.conditional_formatting.add(range, ColorScaleRule(start_type='min', start_color='f8696b',
                                                            mid_type='percent', mid_value=50, mid_color='FFEB84',
                                                            end_type='max', end_color='63BE7B'))
    elif type == 'group':
        ws.conditional_formatting.add(range,
                                      CellIsRule(operator='notEqual', formula=['B2'], stopIfTrue=False,
                                                 fill=PatternFill(bgColor="F8696B")))
    elif type == 'sum':
        ws.conditional_formatting.add(range, ColorScaleRule(start_type='min', start_color='FFFFFF', end_type='max',
                                                            end_color=bgcolors[competence]))
    if type == 'GEK' or type == 'GK' or type == 'K':
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_n,
                                                  formula=text_formula('n', range)))
        ws.conditional_formatting.add(range, Rule(type="containsText", operator='containsText', dxf=style_n,
                                                  formula=text_formula('-', range)))


if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active

    custom_conditional_formatting(ws, 'A2:A5', 'GK')
    custom_conditional_formatting(ws, 'B2:B5', 'EK')
    custom_conditional_formatting(ws, 'C2:C7', 'GEK')

    wb.save("formated.xlsx")
