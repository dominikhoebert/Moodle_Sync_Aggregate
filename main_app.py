import sys
from datetime import datetime
from locale import atof, setlocale, LC_NUMERIC
import json
import os

import pandas as pd
from openpyxl import Workbook, worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QCheckBox, QFileDialog, QScrollArea, QVBoxLayout, \
    QGroupBox
from PyQt5.QtCore import QSettings, QPoint, QSize, Qt

from main_window import Ui_MainWindow
from SettingsDlg import SettingsDlg
from moodle_sync import MoodleSync
from conditional_formating import custom_conditional_formatting
from data_classes import GradeBook, GradePage, Competence, Module
from configparser_module import Config


# Translate .ui to .py
# python -m PyQt5.uic.pyuic -x moodle_sync_aggregate.ui -o main_window.py (mac)
# pyuic5 -o main_window.py .\moodle_sync_aggregate.ui (win)
# pyuic5 -o settings_dialog.py .\settings.ui

# Pyinstaller
# pyinstaller -n Moodle_Sync_Aggregate --onefile --windowed main_app.py


def list_to_float(grade_list):
    """ Function takes a list of values and trys to convert them to floats.
    Returns a list of values which contains str and floats"""
    return_list = []
    for item in grade_list:
        temp_item = item
        try:
            temp_item = float(atof(item))
        except (ValueError, TypeError, AttributeError):
            pass
        return_list.append(temp_item)
    return return_list


def filter_blank(list_to_filter):
    """Filters empty/blank string values from a list of strings
    returns the filtered list"""
    return_list = []
    for i, item in enumerate(list_to_filter):
        if item == '':
            return_list.append(f"Spalte{i}")
        else:
            return_list.append(item)
    return return_list


def get_column_for_module(ws, module):
    """Returns the column letter of the wanted module in a worksheet"""
    for cell in ws[1]:
        if module == cell.value:
            return cell.column_letter


def show_messagebox(message, error=None):
    """Prints a Message in a Messagebox"""
    print(message, error)
    msg_box = QMessageBox(QMessageBox.Information, "Message", message)
    msg_box.exec_()


def replace_grades(grades):
    """replaces long grades with the equivalent shortcut in a grades dataframe
    returns the replaced dataframe"""
    replaces = {"nicht erfüllt": "n", "Nicht erfüllt": "n", "GK vollständig": "GKv", "GK überwiegend": "GKü",
                "EK vollständig": "EKv", "EK überwiegend": "EKü", "vollständig erfüllt": "v",
                "überwiegend erfüllt": "ü"}
    for k, v in replaces.items():
        grades = grades.replace(k, v)
    return grades


def load_student_list(path):
    """ trys to open a studentlist.csv from path. Returns the studentlist as dataframe"""
    try:
        return pd.read_csv(path)
    except FileNotFoundError as e:
        show_messagebox(f"No studentlist at {path}", e)
        return None


def merge_student_list_to_grades(grades, student_list):
    """merges the classes to a gradeslist"""
    if student_list is not None:
        grades = grades.merge(student_list, how='left', left_on='Email', right_on='mail')
        grades = grades.drop(['dn', 'mail', 'sn', 'givenName', 'name', 'accountexpirationdate', 'Email2'], axis=1,
                             errors='ignore')
        grades = grades.rename(columns={'department': 'Klasse'})
    else:
        grades["Klasse"] = ""
    return grades


def open_competence_names_katalog(path):
    """trys to open a competence.json from path. Returns katalog as dict. returns empty dict when failed"""
    try:
        with open(path, 'r') as f:
            competence_names = json.load(f)
    except FileNotFoundError:
        print(f"Competence Names Katalaog at {path} not found.")
        competence_names = {}
    return competence_names


def get_new_worksheet(workbook, title):
    """returns new worksheet, with title"""
    # new workbook already has a worksheet. change title and return it.
    if workbook.sheetnames[0] == 'Sheet':
        ws = workbook.active
        ws.title = title
        return ws
    else:
        return workbook.create_sheet(title)


def grades_page_to_excel_worksheet(grade_page, wb):
    """ converts dataframe to excel worksheet """
    ws = get_new_worksheet(wb, grade_page.name)
    for row in dataframe_to_rows(grade_page.grades, index=False, header=True):
        ws.append(list_to_float(row))
    return ws


def set_column_width(ws):
    col_dims = {'A': 35, 'B': 6, 'C': 8, 'D': 10}
    for col, dim in col_dims.items():
        ws.column_dimensions[col].width = dim
    for i in range(len(col_dims) + 1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 4
    return ws


def create_points_config(column_letter, max_row, ws):
    """ Creates a section at the button end of the table set the needed and max points for the conditional formating"""
    ws[f'A{max_row + 2}'].value = 'Bestehungsgrenze'
    ws[f'A{max_row + 2}'].font = Font(bold=True)
    ws[f'A{max_row + 3}'].value = 'Maximal erreichbar'
    ws[f'A{max_row + 3}'].font = Font(bold=True)
    ws[f'{column_letter}{max_row + 2}'].value = 6
    ws[f'{column_letter}{max_row + 3}'].value = 10
    return ws


class Window(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)

        self.config_path = "config.ini"
        self.config = Config(self.config_path).config['Moodle_Sync_Aggregate']

        self.gradesLayout = QVBoxLayout()
        self.gradesLayout.setAlignment(Qt.AlignTop)
        self.groupbox = QGroupBox('Grades')
        self.groupbox.setLayout(self.gradesLayout)
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidget(self.groupbox)
        self.scroll_area.setWidgetResizable(True)
        self.verticalLayout_2.addWidget(self.scroll_area)

        self.pagesLayout = QVBoxLayout()
        self.pagesLayout.setAlignment(Qt.AlignTop)
        self.groupbox_2 = QGroupBox('Pages')
        self.groupbox_2.setLayout(self.pagesLayout)
        self.scroll_area_2 = QScrollArea()
        self.scroll_area_2.setWidget(self.groupbox_2)
        self.scroll_area_2.setWidgetResizable(True)
        self.verticalLayout_3.addWidget(self.scroll_area_2)

        self.export_pushButton.setEnabled(False)
        self.save_pushButton.setEnabled(False)
        self.merge_pushButton.setEnabled(False)
        self.download_pushButton.setEnabled(False)

        # Event Handlers
        self.courselistWidget.currentItemChanged.connect(self.course_changed)
        self.download_pushButton.clicked.connect(self.download_grades)
        self.export_pushButton.clicked.connect(self.export_grades)
        self.reload_pushButton.clicked.connect(self.download_courses)
        self.actionSettings.triggered.connect(self.open_settings)
        self.save_pushButton.clicked.connect(self.save_grades)
        self.merge_pushButton.clicked.connect(self.merge)

        # Data
        self.current_course = None  # Text
        self.selected_course = None  # Text
        self.courses = None  # Dict Course name:id
        self.modules_checkboxes = None  # List of Checkboxes for Modules
        self.pages_checkboxes = []  # List of Checkboxes for Pages
        self.student_list = None  # Dataframe Name Klasse
        self.competence_catalog = open_competence_names_katalog('modules.json')
        self.grade_book = GradeBook(self.competence_catalog)
        self.current_grades_df = None

        # Config
        self.settings = QSettings(self.config['qsettingscompany'], self.config['qsettingsapplication'])
        # self.settings.clear()
        self.resize(self.settings.value("size", QSize(1000, 800)))
        self.move(self.settings.value("pos", QPoint(50, 50)))
        if self.settings.contains('splitter'):
            self.splitter.restoreState(self.settings.value('splitter'))
        self.use_student_list = self.settings.value("use_student_list", False)
        self.create_competence_columns = self.config.getboolean('competence_columns')
        self.mark_suggestion = self.config.getboolean('marksuggestions')
        self.negative_competences = self.config.getboolean('negative_competences')
        self.competence_counter = self.config.getboolean('competence_counter')
        self.wh_calculation = self.config.getboolean('wh_calculation')
        self.cache_grades = self.config.getboolean('cache_grades')
        self.number_cancel = self.settings.value('number_cancel', 2)
        if self.use_student_list == 'true' or self.use_student_list is True:
            self.use_student_list = True
        else:
            self.use_student_list = False

        # Startup
        self.url = self.config['MoodleURL']
        self.service = self.config['Service_name']
        self.username = self.settings.value("username", None)
        self.password = self.settings.value("password", None)
        self.ldap_username_extension = self.config['Username_extension']
        self.student_list_path = self.settings.value("student_list",
                                                     "~/tgm - Die Schule der Technik/HIT - Abteilung für Informations"
                                                     "technologie - Dokumente/Organisation/Tools/studentlistv2.csv")
        self.ldap_student_list_path = self.config['ldap_student_list_path']
        self.ldap_url = self.config['ldap_url']
        self.moodle = None

        self.login()

    def login(self):
        """login to moodle"""
        if self.username and self.password:
            self.moodle = MoodleSync(self.url, self.username, self.password, self.service)
            self.download_courses()
        else:
            show_messagebox("Moodle Login not defined. Please check Settings.")

    def get_course_id(self, name):
        """returns the moodle id to a course"""
        return self.courses[name]['id']

    def download_courses(self):
        """downloads recent courses"""
        if self.moodle:
            try:
                self.courses = self.moodle.get_recent_courses()
                self.set_courses()
                self.download_pushButton.setEnabled(True)
            except Exception as e:
                show_messagebox("Failed to load courses. Please check Settings.", e)
        else:
            show_messagebox("Moodle URL/Key not defined. Please check Settings.")

    def set_courses(self):
        """ displays in the ui courses """
        self.courselistWidget.blockSignals(True)
        self.courselistWidget.clear()
        for c in self.courses.keys():
            self.courselistWidget.addItem(c)
        self.courselistWidget.blockSignals(False)

    def course_changed(self, course):
        self.selected_course = course.text()

    def merge_group_to_grades(self, grades):
        """ loads and merges the user info (for the groups) for all students in a course """
        user_list = []
        for uid in grades.userid:
            user_list.append({"userid": uid, "courseid": self.get_course_id(self.current_course)})

        user_info = self.moodle.get_student_info(userlist=user_list)
        grades = grades.merge(user_info, how='left', left_on='userid', right_on='id')
        grades = grades.drop(['userid', 'id', 'fullname', 'Email2'], axis=1, errors='ignore')
        grades = grades.rename(columns={'groups': 'Gruppen', 'email': 'Email'})
        return grades

    def download_grades(self):
        """When Download Button pressed.
        loads studentlist, loads grades from selected course, preparing grades, creating checkboxes."""
        # Load Studentlist: first try from ldap then from external path
        if self.student_list is None:
            self.student_list = load_student_list(self.ldap_student_list_path)
            if self.use_student_list and self.student_list is not None:
                self.student_list = load_student_list(self.student_list_path)

        # Check if a course is selected. if not choose first one
        if self.selected_course is None:
            self.courselistWidget.setCurrentRow(0)

        self.current_course = self.selected_course

        files = [f.split(".")[0] for f in os.listdir('data/')]
        if self.cache_grades:
            if self.current_course in files:
                grades = pd.read_csv(f"data/{self.current_course}.csv")
                grades = grades.drop(['Unnamed: 0'], axis=1, errors='ignore')
                print(f"{self.current_course} loaded from file")
            else:
                grades = self.moodle.get_gradereport_of_course(self.get_course_id(self.current_course))
                print(f"{self.current_course} saved to file")
        else:
            grades = self.moodle.get_gradereport_of_course(self.get_course_id(self.current_course))
        grades = self.merge_group_to_grades(grades)
        grades = grades.sort_values(by=['Gruppen', 'Schüler'])

        grades = replace_grades(grades)
        grades = merge_student_list_to_grades(grades, self.student_list)

        grades.columns = filter_blank(grades.columns)
        grades = grades[["Schüler", "Klasse", "Gruppen", "Email"] + list(grades.columns)[1:-3]]

        current_page = GradePage(self.current_course, grades, self.competence_catalog)

        if self.mark_suggestion:
            grades.insert(len(grades.columns), 'Punkte', '=')
            grades.insert(len(grades.columns), 'Notenvorschlag', '=')

        if self.create_competence_columns:
            for competence in current_page.competences:
                grades.insert(len(grades.columns), competence.name, '=')

        if self.negative_competences:
            grades.insert(len(grades.columns), 'Negative Kompetenzen', '=')

        if self.competence_counter:
            grades.insert(len(grades.columns), 'ΣN', '=')
            grades.insert(len(grades.columns), 'ΣGKü', '=')
            grades.insert(len(grades.columns), 'ΣGKv', '=')
            grades.insert(len(grades.columns), 'ΣEKü', '=')
            grades.insert(len(grades.columns), 'ΣEKv', '=')

        if self.wh_calculation:
            grades.insert(len(grades.columns), '∅SMÜ', '=')

        self.current_grades_df = grades

        self.create_modules_list(grades)

    def create_modules_list(self, grades):
        """ creates checkboxes for every module """
        self.groupbox.setTitle("Grades " + self.current_course)
        self.modules_checkboxes = []
        # delete old checkboxes
        for i in reversed(range(self.gradesLayout.count())):
            self.gradesLayout.itemAt(i).widget().setParent(None)

        # create checkboxes
        self.all_none_checkBox = QCheckBox("All/None")
        self.all_none_checkBox.setChecked(True)
        self.gradesLayout.addWidget(self.all_none_checkBox)
        self.all_none_checkBox.stateChanged.connect(self.all_none_checkbox_changed)
        for module in list(grades.columns):
            if module not in ["Schüler", 'Klasse', 'Gruppen', 'Email', 'Punkte']:
                cb = QCheckBox(module, self)
                cb.setChecked(True)
                self.gradesLayout.addWidget(cb)
                self.modules_checkboxes.append(cb)

        self.save_pushButton.setEnabled(True)
        if len(self.grade_book.pages) > 0:
            self.merge_pushButton.setEnabled(True)
        self.export_pushButton.setEnabled(True)

    def remove_columns(self, dataframe):
        """ removes unchecked columns from modules/columns from grades dataframe"""
        columns_to_drop = [cb.text() for cb in self.modules_checkboxes if cb.checkState() == 0]
        return dataframe.drop(columns_to_drop, axis=1)

    def save_grades(self):
        """store grades dataframe in gradebook"""
        if self.grade_book.get_page_from_name(self.current_course) is None:
            self.grade_book.add_page(self.current_course, self.remove_columns(self.current_grades_df))
            self.add_pages_checkbox(self.current_course)
        else:
            show_messagebox(f"{self.current_course} is already saved.")  # TODO override instead of error message

    def add_pages_checkbox(self, name):
        cb = QCheckBox(name, self)
        cb.setChecked(True)
        self.pagesLayout.addWidget(cb)
        self.pages_checkboxes.append(cb)

    def merge(self):
        """ merges current page to selected pages in gradebook """
        r_page = self.remove_columns(self.current_grades_df)
        # get page from gradebook.pages
        l_pages_merge = [cb.text() for cb in self.pages_checkboxes if cb.isChecked()]
        for l_page_name in l_pages_merge:
            l_grades = self.grade_book.get_page_from_name(l_page_name).grades.copy()
            # merge
            l_grades = l_grades.merge(right=r_page, how="left", left_on="Email", right_on="Email")
            # drop columns
            l_grades = l_grades.drop(["Schüler_y", "Klasse_y", "Gruppen_y", "Email_y", "Punkte_y"], axis=1,
                                     errors="ignore")
            n = len([c for c in l_grades.columns if c.startswith("Negative Kompetenzen")])
            l_grades = l_grades.rename(
                columns={'Schüler_x': 'Schüler', 'Klasse_x': 'Klasse', 'Gruppen_x': 'Gruppen', 'Punkte_x': 'Punkte',
                         "Negative Kompetenzen_x": "Negative Kompetenzen",
                         "Negative Kompetenzen_y": f"Negative Kompetenzen{n}"})
            l_grades = l_grades.fillna('')
            # save
            merged_course_name = f"{l_page_name}+"
            self.grade_book.add_page(merged_course_name, l_grades)
            self.add_pages_checkbox(merged_course_name)

    def export_grades(self):
        """ creates a worksheet for every gradepage.
        works through the title name for formating and forumlar creation """
        if len(self.grade_book.pages) == 0:
            self.save_grades()

        # deleting not wanted pages
        delete_cb = []
        for cb in self.pages_checkboxes:
            if cb.checkState() == 0:
                self.grade_book.pages.remove(self.grade_book.get_page_from_name(cb.text()))
                cb.setParent(None)
                delete_cb.append(cb)

        for cb in delete_cb:
            self.pages_checkboxes.remove(cb)

        # creating worksheets and iterating through pages
        wb = Workbook()
        for page_number, page in enumerate(self.grade_book.pages):

            insert_columns = {}
            for index, col in enumerate(page.grades.columns):
                if col.startswith("SYT") and not col.endswith("*"):
                    insert_columns[col] = index

            for ind, col in enumerate(insert_columns):
                index = insert_columns[col]
                page.grades.insert(index + 1 + ind, col + "*", '=')

            ws = grades_page_to_excel_worksheet(page, wb)
            ws = set_column_width(ws)
            ws.freeze_panes = ws['B1']

            # create table
            tab = worksheet.table.Table(displayName=f'Table{page_number}',
                                        ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
            tab.tableStyleInfo = worksheet.table.TableStyleInfo(name="TableStyleMedium1", showRowStripes=True,
                                                                showColumnStripes=False)
            ws.add_table(tab)

            # iterate first row of grades table
            max_row = ws.max_row
            comp_list = []
            wh_letter_list = []
            original_gkp_dict = {}
            gkp_columns_dict = {}
            negative_comp_col = None
            for cell in ws[1]:
                module = str(cell.value)
                cell_range = f"{cell.column_letter}2:{cell.column_letter}{max_row}"

                if module.startswith("GK"):
                    custom_conditional_formatting(ws, cell_range, 'GK')
                    page.get_module_by_name(module).column_letter = cell.column_letter
                elif module.startswith("EK"):
                    custom_conditional_formatting(ws, cell_range, 'EK')
                    page.get_module_by_name(module).column_letter = cell.column_letter
                elif module.startswith("GEK"):
                    custom_conditional_formatting(ws, cell_range, 'GEK')
                    page.get_module_by_name(module).column_letter = cell.column_letter
                elif module.startswith("SYT") or module.startswith("GKP"):
                    if module.endswith("*"):
                        if negative_comp_col is not None:
                            # find Kompetenz Number in module name
                            dot_index = module.find(".")
                            comp_number = module[dot_index - 1:dot_index + 2]
                            # Formular =WENN(ISTZAHL(AK2);AK2;WENN(ISTZAHL(FINDEN("1.2";AE2));"-";""))
                            # add Formular
                            origin_column = original_gkp_dict[module[:-1]]
                            formular = f'=IF(ISNUMBER({origin_column}#),{origin_column}#,IF(ISNUMBER' \
                                       f'(SEARCH("{comp_number}",{negative_comp_col}#)),"-",""))'
                            gkp_columns_dict[comp_number] = cell.column_letter
                            for c_cell in ws[cell.column_letter]:
                                if c_cell.row > 1:
                                    c_cell.value = formular.replace('#', str(c_cell.row))
                            custom_conditional_formatting(ws, cell_range, type='points2',
                                                          start=f'${cell.column_letter}${max_row + 2}',
                                                          end=f'${cell.column_letter}${max_row + 3}')
                            ws = create_points_config(cell.column_letter, max_row, ws)
                    else:
                        ws.column_dimensions[cell.column_letter].hidden = True
                        original_gkp_dict[module] = cell.column_letter

                elif module.startswith("Wiederholung") or module.startswith("SMÜ"):
                    wh_letter_list.append(cell.column_letter)
                    for cx in ws[cell.column_letter]:
                        if cx.value == '-':
                            cx.value = 0.0
                    custom_conditional_formatting(ws, cell_range, type='points',
                                                  start=f'${cell.column_letter}${max_row + 2}',
                                                  end=f'${cell.column_letter}${max_row + 3}')
                    ws = create_points_config(cell.column_letter, max_row, ws)

                # if Kompetenz
                elif len(module) > 1 and module[1] == '.' and module[0].isnumeric() and module[2].isnumeric():
                    comp_list.append(cell.column_letter)

                    modules = []
                    for m in page.get_competence_by_name(module).modules:
                        if m.type in ['G', 'GE']:
                            modules.append(m)

                    formular = ' & ";" & '.join([f"{m.column_letter}#" for m in modules])

                    for c_cell in ws[cell.column_letter]:
                        if c_cell.value == '=':
                            c_cell.value += formular.replace('#', str(c_cell.row))
                    custom_conditional_formatting(ws, cell_range, 'K')

                elif module == 'Punkte' and self.mark_suggestion:
                    formular = ''
                    for m in page.modules:
                        affected_cell = f"SUMPRODUCT(--EXACT({m.column_letter}#"
                        formular += f'{affected_cell},"GKü"))*-1+{affected_cell},"EKü"))+{affected_cell},"EKv"))*2+'
                        if m.type == 'GK':
                            formular += f'{affected_cell},"ü"))*-1+'
                        elif m.type == 'GEK':
                            formular += f'{affected_cell},"ü"))*-1+'
                        elif m.type == 'EK':
                            formular += f'{affected_cell},"ü"))+'
                            formular += f'{affected_cell},"v"))*2+'

                    for c_cell in ws[cell.column_letter]:
                        if c_cell.value == '=':
                            c_cell.value += formular.replace('#', str(c_cell.row))[:-1]
                    custom_conditional_formatting(ws, cell_range, type='scale')

                elif module == 'Notenvorschlag' and self.mark_suggestion:
                    sc = ws.max_column + 3  # start column

                    # create marks table
                    marks_table = [['Note', 'Schlüssel', 'Anz.', 'P', '', 'Komp.', 'Anz.'],
                                   [5, '', '', '', '', 'GK', len(page.get_modules_by_type(['G']))],
                                   [4, 'alle GK mind. ü', f'={get_column_letter(sc + 6)}2+{get_column_letter(sc + 6)}3',
                                    f'={get_column_letter(sc + 2)}3*-1', '*', 'GEK',
                                    len(page.get_modules_by_type(['GE']))],
                                   [3, 'mind. GKv', 6, f'={get_column_letter(sc + 2)}4-{get_column_letter(sc + 2)}3',
                                    '',
                                    'EK', len(page.get_modules_by_type(['E']))],
                                   [2, 'mind. EKü', 6, f'={get_column_letter(sc + 2)}5', '', '', ''],
                                   [1, 'mind. EKv', 6, f'={get_column_letter(sc + 2)}6*2', '', '', '']]

                    # print marks table into worksheet
                    for row_number, rows in enumerate(marks_table):
                        for column_number, cell_value in enumerate(rows):
                            ws[f'{get_column_letter(sc + column_number)}{row_number + 1}'].value = cell_value

                    tab = worksheet.table.Table(displayName=f"KeyTable{page_number}",
                                                ref=f"{get_column_letter(sc)}1:{get_column_letter(sc + 3)}6")
                    tab.tableStyleInfo = worksheet.table.TableStyleInfo(name="TableStyleMedium5", showRowStripes=False,
                                                                        showColumnStripes=False)
                    ws.add_table(tab)

                    tab = worksheet.table.Table(displayName=f"KCTable{page_number}",
                                                ref=f"{get_column_letter(sc + 5)}1:{get_column_letter(sc + 6)}4")
                    tab.tableStyleInfo = worksheet.table.TableStyleInfo(name="TableStyleMedium6", showRowStripes=False,
                                                                        showColumnStripes=False)
                    ws.add_table(tab)

                    marks_table_column_dimensions = [8, 15, 8, 5, 5, 10, 8]
                    for col, dim in enumerate(marks_table_column_dimensions):
                        ws.column_dimensions[get_column_letter(sc + col)].width = dim

                    custom_conditional_formatting(ws, f'{get_column_letter(sc)}2:{get_column_letter(sc)}6',
                                                  type='marks')

                    mcl = f'${get_column_letter(sc)}$'  # mark_column_letter
                    kpcl = f'${get_column_letter(sc + 3)}$'  # key_points_column_letter

                    formular_string = '=_xlfn.IFS(SUMPRODUCT(--ISNUMBER(FIND({"n";"-"},'

                    for c_cell in ws[cell.column_letter]:
                        if c_cell.value == '=':
                            c_cell.font = Font(bold=True)
                            pcc = f'{get_column_letter(c_cell.column - 1)}{c_cell.row}'  # points_cell_coordinate
                            cf = ' & '.join(
                                [f'{module.column_letter}{c_cell.row}' for module in page.get_modules_by_type(['G',
                                                                                                               'GE'])])
                            c_cell.value = formular_string + cf + f')))>0,{mcl}2,{pcc}>={kpcl}6,{mcl}6,{pcc}>={kpcl}5,' \
                                                                  f'{mcl}5,{pcc}>={kpcl}4,{mcl}4,{pcc}>={kpcl}3,{mcl}3)'
                    custom_conditional_formatting(ws, cell_range, type='marks')

                elif module == 'Gruppen':
                    custom_conditional_formatting(ws, cell_range, type='group')

                elif module.startswith('Negative Kompetenzen') and self.negative_competences:
                    if module[-1].isnumeric():
                        # formular: =WENN(ODER(AO2="-";AO2<AO$80);"1.1;";"")&WENN(ODER(AP2="-";AP2<AP$80);"1.2;";"")
                        formular = '='
                        for comp_number, col_letter in gkp_columns_dict.items():
                            formular += f'IF(OR({col_letter}#="-",{col_letter}#<{col_letter}${max_row + 2}),"' \
                                        f'{comp_number};","")&'
                        for c_cell in ws[cell.column_letter]:
                            if c_cell.row > 1:
                                c_cell.value = formular.replace('#', str(c_cell.row))[:-1]
                    else:
                        formular_parts = []
                        for comp_letter in comp_list:
                            comp_number = ws[f'{comp_letter}1'].value[:3]
                            formular_parts.append('IF(SUMPRODUCT(--ISNUMBER(FIND({"n";"-"},' +
                                                  f'{comp_letter}#)))>0,"{comp_number};","")')
                        formular = " & ".join(formular_parts)

                        for c_cell in ws[cell.column_letter]:
                            if c_cell.value == '=':
                                c_cell.value += formular.replace('#', str(c_cell.row))
                    ws.column_dimensions[cell.column_letter].width = 14
                    negative_comp_col = cell.column_letter

                elif module in ['ΣN', 'ΣGKü', 'ΣGKv', 'ΣEKü', 'ΣEKv'] and self.competence_counter:
                    custom_conditional_formatting(ws, cell_range, type='sum', competence=module[1:])
                    col_letters = []
                    search_for = []
                    if module == 'ΣN' or module == 'ΣGKü' or module == 'ΣGKv':
                        col_letters.extend([m.column_letter for m in page.get_modules_by_type(['G', 'GE'])])
                        if module == 'ΣN':
                            search_for = ['n', '-']
                        if module == 'ΣGKü':
                            search_for = ['ü', 'GKü']
                        elif module == 'ΣGKv':
                            search_for = ['v', 'GKv']
                    elif module == 'ΣEKü' or module == 'ΣEKv':
                        col_letters.extend([m.column_letter for m in page.get_modules_by_type(['E', 'GE'])])
                        if module == 'ΣEKü':
                            search_for = ['ü', 'EKü']
                        elif module == 'ΣEKv':
                            search_for = ['v', 'EKv']
                    formular_parts = []
                    for letter in col_letters:
                        for sf in search_for:
                            formular_parts.append(f'SUMPRODUCT(--EXACT({letter}#,"{sf}"))')
                    formular = "+".join(formular_parts)
                    for c_cell in ws[cell.column_letter]:
                        if c_cell.value == '=':
                            c_cell.value += formular.replace('#', str(c_cell.row))

                elif module == '∅SMÜ' and self.wh_calculation:
                    # =(SUM(F2:L2)-SMALL(F2:L2; 1)-SMALL(F2:L2; 2))/5
                    formular_range = "#,".join(wh_letter_list) + "#"
                    formular = f"(SUM({formular_range})"
                    for i in range(self.number_cancel):
                        formular += f"-SMALL(({formular_range}), {i + 1})"
                    formular += f")/{len(wh_letter_list) - self.number_cancel}"
                    for c_cell in ws[cell.column_letter]:
                        if c_cell.value == '=':
                            c_cell.value += formular.replace('#', str(c_cell.row))
                    custom_conditional_formatting(ws, cell_range, type='points',
                                                  start=f'${cell.column_letter}${max_row + 2}',
                                                  end=f'${cell.column_letter}${max_row + 3}')
                    ws = create_points_config(cell.column_letter, max_row, ws)

        # prepare for save to file
        directory = self.settings.value('dir', "")
        ct = datetime.now()
        filename = f"{directory}/{ct.year}{str(ct.month).zfill(2)}{str(ct.day).zfill(2)}_Noten"
        if len(self.grade_book.pages) == 1:
            filename += '_' + self.current_course

        file, _ = QFileDialog.getSaveFileName(self, "Export Grades", filename, "Excel files (*.xlsx)")
        if file:
            wb.save(file)
            self.settings.setValue("dir", file[:file.rfind("/")])

    def open_settings(self):
        config_text = f'Config file at {self.config_path}:\nMoodleURL: {self.url}\n' \
                      f'Service_name: {self.service}\nUsername_extension: {self.ldap_username_extension}\n' \
                      f'cache_grades: {self.cache_grades}\nmarksuggestions: {self.mark_suggestion}\n' \
                      f'competence_columns: {self.create_competence_columns}\nnegative_competences: ' \
                      f'{self.negative_competences}\ncompetence_counter: {self.competence_counter}\n' \
                      f'wh_calculation: {self.wh_calculation}\nldap_student_list_path: {self.ldap_student_list_path}\n'\
                      f'ldap_url: {self.ldap_url}'

        settings = SettingsDlg(self.username, self.password, self.use_student_list, self.student_list_path,
                               self.number_cancel, ldap_url=self.ldap_url, filename=self.ldap_student_list_path,
                               config_text=config_text, parent=self)
        if settings.exec():
            self.username = settings.ui.username_lineEdit.text()
            self.password = settings.ui.password_lineEdit.text()
            self.use_student_list = settings.ui.checkBox.isChecked()
            self.student_list_path = settings.ui.studentlist_lineEdit.text()
            self.number_cancel = settings.ui.cancle_number_spinBox.value()

            self.settings.setValue("username", self.username)
            self.settings.setValue("password", self.password)
            self.settings.setValue("use_student_list", self.use_student_list)
            self.settings.setValue("student_list", self.student_list_path)
            self.settings.setValue("number_cancel", self.number_cancel)
        self.login()

    def all_none_checkbox_changed(self):
        for cb in self.modules_checkboxes:
            cb.setChecked(self.all_none_checkBox.checkState())

    def closeEvent(self, e):
        self.settings.setValue("size", self.size())
        self.settings.setValue("pos", self.pos())
        self.settings.setValue("splitter", self.splitter.saveState())
        e.accept()


if __name__ == '__main__':
    setlocale(LC_NUMERIC, 'de_DE')
    app = QApplication(sys.argv)
    win = Window()
    win.show()
    sys.exit(app.exec())
