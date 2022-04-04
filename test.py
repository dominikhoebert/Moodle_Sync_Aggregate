import json
import pandas as pd
from openpyxl import Workbook, worksheet

if __name__ == '__main__':
    with open('data/credentials.json', 'r') as f:
        data = json.load(f)
    url = data['url']
    key = data['key']
    username = data['username']
    password = data['password']
    print(url, key, username, password)

    df1 = pd.DataFrame([['a', 'b'], ['x', 'y']], columns=['first', 'sec'])
    df2 = pd.DataFrame([['c', 'd'], ['z', 'z']], columns=['first', 'sec'])

    wb = Workbook()
    for p in ['1a', '2a', '3a']:
        if wb.sheetnames[0] == 'Sheet':
            ws = wb.active
            ws.title = p
        else:
            ws = wb.create_sheet(p)

    print('final:', wb.sheetnames)
